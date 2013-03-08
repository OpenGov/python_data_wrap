from openpyxl import load_workbook
import xlrd
import re,csv,os

# Used throughout -- never changed
xlsxExtRegex = re.compile(r'(\.xlsx)\s*$')
xlsExtRegex = re.compile(r'(\.xls)\s*$')
csvExtRegex = re.compile(r'(\.csv)\s*$')

'''
Loads an arbitrary file type (xlsx, xls, or csv like) and returns
a list of 2D tables. For csv files this will be a list of one table,
but excel formats can have many tables/worksheets.

@author Matt Seal
'''
def read(filename):
    if re.search(xlsxExtRegex, filename):
        return getDataXlsx(filename)
    elif re.search(xlsExtRegex, filename):
        return getDataXls(filename)
    elif re.search(csvExtRegex, filename):
        return getDataCsv(filename)
    else:
        try:
            return getDataCsv(filename)
        except:
            raise ValueError("Unable to load file '"+filename+"' as csv")
        
'''
Gets new version excel data. This will not load old '.xls' files.

@author Joe Maguire
@author Matt Seal
'''
def getDataXlsx(filename):
    wb = load_workbook(filename)
    data = []
    
    for ws in wb.worksheets:
        dimension = ws.calculate_dimension()
        ws_data = []
        
        for row in ws.range(dimension):
            data_row = []
           
            for cell in row:
                data_row.append(cell.value)
            ws_data.append(data_row)
        
        data.append(ws_data)
    
    return data

'''
Gets old version excel data. This will not load new '.xlsx' files.

@author Joe Maguire
@author Matt Seal
'''
def getDataXls(filename):
    def tupledateToIsoDate(tupledate):
        """
        Turns a gregorian (year, month, day, hour, minute, nearest_second) into a
        standard YYYY-MM-DDTHH:MM:SS ISO date.  If the date part is all zeros, it's
        assumed to be a time; if the time part is all zeros it's assumed to be a date;
        if all of it is zeros it's taken to be a time, specifically 00:00:00 (midnight).
    
        Note that datetimes of midnight will come back as date-only strings.  A date
        of month=0 and day=0 is meaningless, so that part of the coercion is safe.
        For more on the hairy nature of Excel date/times see http://www.lexicon.net/sjmachin/xlrd.html
        """
        (y,m,d, hh,mm,ss) = tupledate
        nonzero = lambda n: n!=0
        date = "%04d-%02d-%02d"  % (y,m,d)    if filter(nonzero, (y,m,d))                else ''
        time = "T%02d:%02d:%02d" % (hh,mm,ss) if filter(nonzero, (hh,mm,ss)) or not date else ''
        return date+time

    def formatExcelVal(book, val_type, value, wanttupledate):
        """ Clean up the incoming excel data """
        ##  Data val_type Codes:
        ##  EMPTY   0
        ##  TEXT    1 a Unicode string 
        ##  NUMBER  2 float 
        ##  DATE    3 float 
        ##  BOOLEAN 4 int; 1 means TRUE, 0 means FALSE 
        ##  ERROR   5 
        if   val_type == 2: # TEXT
            if value == int(value): value = int(value)
        elif val_type == 3: # NUMBER
            datetuple = xlrd.xldate_as_tuple(value, book.datemode)
            value = datetuple if wanttupledate else tupledateToIsoDate(datetuple)
        elif val_type == 5: # ERROR
            value = xlrd.error_text_from_code[value]
        return value
    
    def xlrdXlsToArray(filename):
        """ Returns a list of sheets; each sheet is a dict containing
        * sheet_name: unicode string naming that sheet
        * sheet_data: 2-D table holding the converted cells of that sheet
        """    
        book       = xlrd.open_workbook(filename)
        sheets     = []
        formatter  = lambda(t,v): formatExcelVal(book,t,v,False)
        
        for sheet_name in book.sheet_names():
            raw_sheet = book.sheet_by_name(sheet_name)
            data      = []
            for row in range(raw_sheet.nrows):
                (types, values) = (raw_sheet.row_types(row), raw_sheet.row_values(row))
                data.append(map(formatter, zip(types, values)))
            sheets.append({'sheet_name': sheet_name, 'sheet_data': data})
        return sheets
    
    data = []
    
    for ws in xlrdXlsToArray(filename):
        data.append(ws['sheet_data'])
    return data

'''
Gets good old csv data from a file.

@author Matt Seal
'''
def getDataCsv(filename, loadAsUnicode=True):
    table = []
    
    with open(filename, "rb") as file:
        csvfile = csv.reader(file, dialect=csv.excel)
        for line in csvfile:
            if loadAsUnicode:
                table.append([unicode(cell, 'utf-8') for cell in line])
            else:
                table.append(line)
            
    return [table]

'''
Writes 2D tables to file.

@param data 2D list of tables/worksheets
@param filename Name of the output file (determines type)
@author Matt Seal
'''
def write(data, filename):
    if re.search(xlsxExtRegex, filename):
        return writeXlsx(data, filename)
    elif re.search(xlsExtRegex, filename):
        return writeXls(data, filename)
    elif re.search(csvExtRegex, filename):
        return writeCsv(data, filename)
    else:
        return writeCsv(data, filename)
      
'''
Writes out to new excel format.

@param data 2D list of tables/worksheets
@param filename Name of the output file
@author Matt Seal
'''  
def writeXlsx(data, filename):
    raise NotImplementedError("Xlsx writing not implemented")

'''
Writes out to old excel format.

@param data 2D list of tables/worksheets
@param filename Name of the output file
@author Matt Seal
''' 
def writeXls(data, filename):
    raise NotImplementedError("Xls writing not implemented")

'''
Writes out to csv format.

@param data 2D list of tables/worksheets
@param filename Name of the output file
@author Matt Seal
'''    
def writeCsv(data, filename):
    nameExtension = len(data) > 1
    root, ext = os.path.splitext(filename)
    
    for i, sheet in enumerate(data):
        fname = filename if not nameExtension else root+"_"+str(i)+ext
        with open(fname, "wb") as file:
            csvfile = csv.writer(file)
            for line in sheet:
                csvfile.writerow(line)
    